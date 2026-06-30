"""
Isolates classification logic and boundary detection.
Implements custom formula hypervisor enforcing intra-row recalculations,
headless column pruning, strict AMT calculations via Price Groups,
and contiguous serial re-indexing.
"""

import logging
import re

from bom_dic.bom_split.config_split import (
    END_MARKER,
    HEADER_ROW,
    PANEL_ROW,
    PANEL_START_COL,
    PRICE_AMT,
    PRICE_GROUPS,
    PRICE_RATE,
)
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet


def _is_empty(val: object) -> bool:
    if val is None:
        return True
    return str(val).strip().upper() in ("", "0", "0.0", "NONE", "NULL", "-")


def _is_numeric(val: object) -> bool:
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        float(str(val))
        return True
    except ValueError:
        return False


def _sanitize_sheet_title(title: str) -> str:
    safe_title = re.sub(r"[\\/*?:\[\]]", "_", title)
    return safe_title[:31]


def _get_shifted_merges(ws: Worksheet, cols_to_delete: list[int]) -> list[CellRange]:
    new_ranges = []
    for mr in list(ws.merged_cells.ranges):
        shift_min = sum(1 for c in cols_to_delete if c < mr.min_col)
        shift_max = sum(1 for c in cols_to_delete if c < mr.max_col)

        new_min_col = mr.min_col - shift_min
        new_max_col = mr.max_col - shift_max

        if new_min_col <= new_max_col:
            new_mr = CellRange(
                min_col=new_min_col,
                max_col=new_max_col,
                min_row=mr.min_row,
                max_row=mr.max_row,
            )
            new_ranges.append(new_mr)
    return new_ranges


def _shift_formulas_and_unhide(ws: Worksheet, cols_to_delete: list[int]) -> None:
    shift_map = {}
    for col_idx in range(1, 2000):
        shift = sum(1 for c in cols_to_delete if c < col_idx)
        shift_map[col_idx] = col_idx - shift

    pattern = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")

    for row in ws.iter_rows():
        current_row = row[0].row

        def replacer(match: re.Match[str], row_idx: int | None = current_row) -> str:
            abs_col = match.group(1)
            col_letters = match.group(2)
            abs_row = match.group(3)

            try:
                col_idx = column_index_from_string(col_letters)
            except ValueError:
                return match.group(0)

            if col_idx in cols_to_delete:
                return "#REF!"

            new_col_idx = shift_map.get(col_idx, col_idx)
            new_col_letters = get_column_letter(new_col_idx)

            return f"{abs_col}{new_col_letters}{abs_row}{row_idx}"

        for cell in row:
            if not isinstance(cell, Cell):
                continue

            if isinstance(cell.value, str):
                if cell.value.startswith("FORMULA_HIDE="):
                    orig_formula = cell.value.replace("FORMULA_HIDE", "", 1)
                    cell.value = pattern.sub(replacer, orig_formula)
                elif cell.value.startswith("'="):
                    orig_formula = cell.value.replace("'", "", 1)
                    cell.value = pattern.sub(replacer, orig_formula)


def _fix_amt_formulas(ws: Worksheet) -> None:
    """
    Dynamically maps PRICE_RATE * QTY into PRICE_AMT cells.
    Safeguards footer rows from being overridden by zero-value formulas,
    and skips formula calculation entirely if the PRICE_RATE is empty.
    """
    # 1. Locate the surviving QTY column
    qty_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=HEADER_ROW, column=c).value or "").strip().upper() == "QTY":
            qty_col = c
            break

    if not qty_col:
        return

    qty_col_letter = get_column_letter(qty_col)

    # 2. Locate and Pair up Rate / Amt Columns dynamically based on Config
    rate_cols = []
    amt_cols = []
    for c in range(1, ws.max_column + 1):
        header = str(ws.cell(row=HEADER_ROW, column=c).value or "").strip().upper()
        if header == PRICE_RATE.upper():
            rate_cols.append(c)
        elif header == PRICE_AMT.upper():
            amt_cols.append(c)

    pairs = []
    for amt_c in amt_cols:
        left_rates = [rc for rc in rate_cols if rc < amt_c]
        if left_rates:
            pairs.append((left_rates[-1], amt_c))
        else:
            pairs.append((amt_c - 1, amt_c))

    # 3. Apply the specific PRICE_RATE * QTY formula
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        qty_val = ws.cell(row=r, column=qty_col).value
        sno_val = ws.cell(row=r, column=1).value

        # Only overwrite if it is a standard row to protect Footer summary lines
        if not _is_empty(qty_val) or _is_numeric(sno_val):
            for rate_c, amt_c in pairs:
                rate_val = ws.cell(row=r, column=rate_c).value
                target_cell = ws.cell(row=r, column=amt_c)

                if isinstance(target_cell, Cell):
                    if _is_empty(rate_val):
                        target_cell.value = None
                    else:
                        rate_col_letter = get_column_letter(rate_c)
                        target_cell.value = f"={rate_col_letter}{r}*{qty_col_letter}{r}"


def _clear_headless_columns(ws: Worksheet) -> None:
    for c in range(1, ws.max_column + 1):
        has_header = False
        for r in range(1, 6):
            if not _is_empty(ws.cell(row=r, column=c).value):
                has_header = True
                break

        if not has_header:
            for r in range(1, ws.max_row + 1):
                target_cell = ws.cell(row=r, column=c)
                if isinstance(target_cell, Cell):
                    target_cell.value = None


def _reindex_serial_numbers(ws: Worksheet, sno_col: int) -> None:
    current_sno = 1
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        target_cell = ws.cell(row=r, column=sno_col)
        if not isinstance(target_cell, Cell):
            continue

        sno_val = target_cell.value
        if _is_numeric(sno_val):
            target_cell.value = current_sno
            current_sno += 1


def detect_columns(
    ws: Worksheet, split_sub: bool, logger: logging.Logger
) -> tuple[dict[str, list[int]], list[int], dict[str, int]]:
    panel_cols: dict[str, list[int]] = {}
    all_panel_cols: list[int] = []
    header_map: dict[str, int] = {}

    end_col: int = -1
    for col in range(PANEL_START_COL, ws.max_column + 1):
        val1 = str(ws.cell(row=1, column=col).value or "").strip().upper()
        val2 = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip().upper()
        if END_MARKER.upper() in val1 or END_MARKER.upper() in val2:
            end_col = col
            break

    if end_col == -1:
        logger.warning(f"Boundary marker '{END_MARKER}' not found. Scanning to max.")
        end_col = ws.max_column + 1

    for col in range(PANEL_START_COL, end_col):
        raw_val = ws.cell(row=PANEL_ROW, column=col).value
        val = str(raw_val or "").strip()
        if val:
            all_panel_cols.append(col)
            prefix = val if split_sub else val.split("-")[0].strip()
            if prefix not in panel_cols:
                panel_cols[prefix] = []
            panel_cols[prefix].append(col)

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=HEADER_ROW, column=col).value
        if val:
            header_map[str(val).strip().upper()] = col

    return panel_cols, all_panel_cols, header_map


def classify_panels(
    wb: Workbook,
    ws: Worksheet,
    panel_cols: dict[str, list[int]],
    all_panel_cols: list[int],
    header_map: dict[str, int],
    logger: logging.Logger,
) -> int:
    sheets_created: int = 0
    sno_col: int = header_map.get("SNO", 1)
    desc_col: int = header_map.get("DESCRIPTION", 2)
    cat_col: int = header_map.get("CAT NO.", 6)

    for prefix, prefix_panel_cols in panel_cols.items():
        logger.info(f"Generating isolated structure for series: {prefix}")

        ws_new: Worksheet = wb.copy_worksheet(ws)
        ws_new.title = _sanitize_sheet_title(prefix)

        data_rows_to_delete: set[int] = set()

        for r in range(HEADER_ROW + 1, ws_new.max_row + 1):
            sno_val = ws_new.cell(row=r, column=sno_col).value
            cat_val = ws_new.cell(row=r, column=cat_col).value

            is_data_row = _is_numeric(sno_val) or not _is_empty(cat_val)

            has_prefix_qty = any(
                not _is_empty(ws_new.cell(row=r, column=c).value)
                and ws_new.cell(row=r, column=c).value != 0
                for c in prefix_panel_cols
            )

            if is_data_row:
                if not has_prefix_qty:
                    data_rows_to_delete.add(r)
            else:
                has_label = not _is_empty(ws_new.cell(row=r, column=desc_col).value)
                if not has_prefix_qty and not has_label:
                    data_rows_to_delete.add(r)

        rows_to_delete: list[int] = sorted(data_rows_to_delete, reverse=True)
        for r in rows_to_delete:
            ws_new.delete_rows(r)

        protected_cols = set()
        for c in range(1, ws_new.max_column + 1):
            v1 = str(ws_new.cell(row=1, column=c).value or "").strip().upper()
            v2 = str(ws_new.cell(row=PANEL_ROW, column=c).value or "").strip().upper()
            v3 = str(ws_new.cell(row=HEADER_ROW, column=c).value or "").strip().upper()

            is_protected = False
            if v3 in (PRICE_RATE.upper(), PRICE_AMT.upper()):
                is_protected = True

            for pg in PRICE_GROUPS:
                if pg.upper() in v1 or pg.upper() in v2:
                    is_protected = True
                    break

            if is_protected:
                protected_cols.add(c)

        cols_to_delete: list[int] = sorted(
            [
                c
                for c in all_panel_cols
                if c not in prefix_panel_cols and c not in protected_cols
            ],
            reverse=True,
        )

        for row in ws_new.iter_rows():
            for cell in row:
                if not isinstance(cell, Cell):
                    continue

                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = "FORMULA_HIDE" + cell.value

        shifted_merges = _get_shifted_merges(ws_new, cols_to_delete)
        # pyrefly: ignore [implicit-any-empty-container]
        ws_new.merged_cells.ranges = []

        for c in cols_to_delete:
            ws_new.delete_cols(c)

        target_qty_col = 6
        if prefix_panel_cols:
            first_panel_col = prefix_panel_cols[0]
            shift = sum(1 for c in cols_to_delete if c < first_panel_col)
            target_qty_col = first_panel_col - shift

        actual_desc_col = desc_col - sum(1 for c in cols_to_delete if c < desc_col)
        actual_sno_col = sno_col - sum(1 for c in cols_to_delete if c < sno_col)
        actual_cat_col = cat_col - sum(1 for c in cols_to_delete if c < cat_col)

        for r in range(HEADER_ROW + 1, ws_new.max_row + 1):
            sno_val = ws_new.cell(row=r, column=actual_sno_col).value
            cat_val = ws_new.cell(row=r, column=actual_cat_col).value

            is_data_row = _is_numeric(sno_val) or not _is_empty(cat_val)
            if not is_data_row:
                label_cell = ws_new.cell(row=r, column=actual_desc_col)
                label_val = label_cell.value

                if label_val is not None and str(label_val).strip() != "" and actual_desc_col != target_qty_col - 1:  # noqa: E501
                    target_label_cell = ws_new.cell(
                        row=r, column=target_qty_col - 1
                    )
                    if isinstance(target_label_cell, Cell):
                        target_label_cell.value = label_val
                    if isinstance(label_cell, Cell):
                        label_cell.value = None

        _shift_formulas_and_unhide(ws_new, cols_to_delete)

        _fix_amt_formulas(ws_new)

        for mr in shifted_merges:
            ws_new.merged_cells.add(mr)

        _clear_headless_columns(ws_new)

        _reindex_serial_numbers(ws_new, actual_sno_col)

        sheets_created += 1

    return sheets_created
