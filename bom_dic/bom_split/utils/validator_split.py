"""
Post-processing cryptograph-style numeric verification.
Calculates the dynamic termination boundary (Option B) natively within the
baseline matrix to evaluate strict apples-to-apples line item fidelity.
"""

import logging
from contextlib import suppress
from decimal import Decimal, InvalidOperation

from bom_dic.bom_split.config_split import END_MARKER, HEADER_ROW, PANEL_ROW
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def _is_empty(val: object) -> bool:
    """Evaluates if a cell value constitutes an empty or null equivalent state."""
    if val is None:
        return True
    return str(val).strip().upper() in ("", "0", "0.0", "NONE", "NULL", "-")


def _find_cutoff_row(ws: Worksheet) -> int:
    """
    Scans for the termination boundary (two consecutive rows lacking
    Description and Cat No).
    Returns the integer index of the final valid row before the trailing summary.
    """
    desc_col = -1
    cat_col = -1

    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=HEADER_ROW, column=c).value or "").strip().upper()
        if val == "DESCRIPTION":
            desc_col = c
        elif val == "CAT NO.":
            cat_col = c

    if desc_col == -1 or cat_col == -1:
        return ws.max_row

    consecutive_blanks = 0
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        desc_val = ws.cell(row=r, column=desc_col).value
        cat_val = ws.cell(row=r, column=cat_col).value

        if _is_empty(desc_val) and _is_empty(cat_val):
            consecutive_blanks += 1
            if consecutive_blanks == 2:
                return r - 2
        else:
            consecutive_blanks = 0

    return ws.max_row


def _extract_sum(ws: Worksheet, col_idx: int, cutoff_row: int) -> Decimal:
    """
    Aggregates absolute columnar data utilizing exact arithmetic up to
    the cutoff boundary.
    """
    total = Decimal("0.0")
    for r in range(HEADER_ROW + 1, cutoff_row + 1):
        cell_val = ws.cell(row=r, column=col_idx).value
        if cell_val is not None:
            if isinstance(cell_val, str) and str(cell_val).startswith("="):
                continue
            with suppress(InvalidOperation):
                total += Decimal(str(cell_val).strip())
    return total


def validate_outputs(
    wb_master: Workbook, wb_out: Workbook, logger: logging.Logger
) -> bool:
    """
    Enforces a strict equivalence check between the sanitized source numeric footprint
    and the compiled output to guarantee structural and financial fidelity.
    """
    logger.info("Initiating precision numeric validation protocol.")
    passed: bool = True

    master_sums: dict[str, Decimal] = {}

    for sheet_name in wb_master.sheetnames:
        ws = wb_master[sheet_name]
        cutoff_row = _find_cutoff_row(ws)

        for col in range(6, ws.max_column + 1):
            val = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip().upper()
            if END_MARKER in val:
                break
            panel_name = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip()
            if panel_name:
                master_sums[panel_name] = master_sums.get(
                    panel_name, Decimal("0.0")
                ) + _extract_sum(ws, col, cutoff_row)

    output_sums: dict[str, Decimal] = {}
    for sheet_name in wb_out.sheetnames:
        ws = wb_out[sheet_name]
        cutoff_row = ws.max_row

        for col in range(6, ws.max_column + 1):
            val = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip().upper()
            if END_MARKER in val:
                break
            panel_name = str(ws.cell(row=PANEL_ROW, column=col).value or "").strip()
            if panel_name in master_sums:
                output_sums[panel_name] = output_sums.get(
                    panel_name, Decimal("0.0")
                ) + _extract_sum(ws, col, cutoff_row)

    for panel_name, expected in master_sums.items():
        actual = output_sums.get(panel_name, Decimal("0.0"))
        if expected != actual:
            logger.error(
                f"Fidelity failure on [{panel_name}]. "
                f"Baseline={expected}, Compiled={actual}."
            )
            passed = False

    if passed:
        logger.info(
            "Validation complete: 100% equivalence achieved against "
            "sanitized baseline matrix."
        )

    return passed
