"""
Handles workbook ingestion, applies resource limits via bounds checking,
and routes structural data to the classifier.
"""

import copy as cp
import logging
from pathlib import Path

import openpyxl
from bom_dic.bom_split.config_split import MAX_COLS, MAX_ROWS, OUTPUT_DIR, OUTPUT_FILE
from bom_dic.bom_split.utils.classifier_split import classify_panels, detect_columns
from bom_dic.bom_split.utils.exporter_split import export_exceptions
from bom_dic.bom_split.utils.validator_split import validate_outputs
from openpyxl.worksheet.worksheet import Worksheet


def _verify_boundaries(file_path: Path, logger: logging.Logger) -> bool:
    logger.info(f"Executing boundary verification matrix scan on {file_path.name}...")
    wb_ro = openpyxl.load_workbook(file_path, read_only=True)

    for sheet_name in wb_ro.sheetnames:
        ws_ro = wb_ro[sheet_name]
        for row_count, _ in enumerate(ws_ro.iter_rows(), start=1):
            if row_count > MAX_ROWS:
                logger.error(f"Row limit breached in {sheet_name}. Terminating.")
                return False

        if ws_ro.max_column and ws_ro.max_column > MAX_COLS:
            logger.error(f"Column limit breached in {sheet_name}. Terminating.")
            return False

    wb_ro.close()
    return True


def _copy_sheet_across(ws_src: Worksheet, wb_dest: openpyxl.Workbook) -> None:
    """Deep copies an isolated worksheet across workbooks to create a single output."""
    title = ws_src.title
    base_title = title
    counter = 1
    # Prevent naming collisions if different files have identical panel names
    while title in wb_dest.sheetnames:
        title = f"{base_title}_{counter}"
        counter += 1

    ws_dest = wb_dest.create_sheet(title=title)

    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dest.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cp.copy(cell.font)
                new_cell.border = cp.copy(cell.border)
                new_cell.fill = cp.copy(cell.fill)
                new_cell.number_format = cp.copy(cell.number_format)
                new_cell.protection = cp.copy(cell.protection)
                new_cell.alignment = cp.copy(cell.alignment)

    if ws_src.merged_cells:
        for merge in ws_src.merged_cells.ranges:
            ws_dest.merged_cells.add(cp.copy(merge))

    for col_letter, col_dim in ws_src.column_dimensions.items():
        ws_dest.column_dimensions[col_letter].width = col_dim.width
    for row_idx, row_dim in ws_src.row_dimensions.items():
        ws_dest.row_dimensions[row_idx].height = row_dim.height


def process_workbooks(
    file_paths: list[Path], split_sub: bool, logger: logging.Logger
) -> bool:
    wb_out = openpyxl.Workbook()
    # Remove the default 'Sheet' created upon instantiation
    if wb_out.active:
        wb_out.remove(wb_out.active)

    total_panels_created = 0
    all_exceptions: list[dict[str, str]] = []

    for file_path in file_paths:
        if not _verify_boundaries(file_path, logger):
            continue

        logger.info(f"Ingesting target workbook: {file_path.name}")
        wb = openpyxl.load_workbook(file_path)
        wb_master = openpyxl.load_workbook(file_path, data_only=True)

        master_sheet_names = wb.sheetnames

        for sheet_name in master_sheet_names:
            ws = wb[sheet_name]
            logger.info(f"Targeting matrix: {sheet_name}")

            panel_cols, all_panel_cols, header_map = detect_columns(
                ws, split_sub, logger
            )  # noqa: E501

            if not panel_cols:
                logger.warning(f"No series detected in {sheet_name}.")
                all_exceptions.append(
                    {
                        "sheet": f"{file_path.name} - {sheet_name}",
                        "issue": "Series Void",
                        "description": "Row boundary scan returned no valid nomenclature.",  # noqa: E501
                    }
                )
            else:
                count: int = classify_panels(
                    wb, ws, panel_cols, all_panel_cols, header_map, logger
                )
                total_panels_created += count

            del wb[sheet_name]

        # Validate immediately before copying
        validate_outputs(wb_master, wb, logger)
        wb_master.close()

        # Transfer all processed sheets into the single final workbook
        for ws in wb.worksheets:
            _copy_sheet_across(ws, wb_out)

        wb.close()

    if total_panels_created == 0:
        logger.error("Architecture operation failed. Zero distinct panels generated.")
        return False

    out_path: Path = OUTPUT_DIR / OUTPUT_FILE
    wb_out.save(out_path)
    logger.info(f"Unified Workbook compiled and locked at {out_path}")

    export_exceptions(all_exceptions, logger)
    return True
