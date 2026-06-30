"""
Loads a .xlsx workbook and extracts BoM rows from every sheet.
Visual gaps are bypassed via bounded iteration based on max_row.
"""

import logging
from typing import Any

import openpyxl
from bom_dic.bom_merge.config_merge import HEADER_ROW, STATIC_COLS
from bom_dic.bom_merge.utils.panel_detector_merge import detect_panels
from bom_dic.bom_merge.utils.row_cleaner_merge import clean_row
from openpyxl.worksheet.worksheet import Worksheet

SRNO_COL = 1


def _build_header_map(ws: Worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    col = 1
    while True:
        cell = ws.cell(row=HEADER_ROW, column=col)
        if col > 150:
            break
        raw = cell.value
        if raw is not None:
            normalised = str(raw).strip().upper()
            if normalised in STATIC_COLS:
                header_map[normalised] = col
        col += 1
    return header_map


def _verify_srno_header(ws: Worksheet) -> bool:
    val = ws.cell(row=HEADER_ROW, column=SRNO_COL).value
    if val is None:
        return False
    return str(val).strip().upper() == "SNO"


def read_workbook(
    file_path: str,
    logger: logging.Logger,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets_data: list[dict[str, Any]] = []
    exceptions: list[dict[str, Any]] = []

    def _is_empty(val: object) -> bool:
        if val is None:
            return True
        text = str(val).strip().upper()
        return text in ("", "0", "0.0", "NONE", "NULL", "-")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_category = "Uncategorized"

        panels = detect_panels(ws)
        if not panels:
            logger.warning(
                f"Sheet '{sheet_name}': no panel names in row 1 from col G. Skipping."
            )
            exceptions.append(
                {
                    "sheet": sheet_name,
                    "row": 1,
                    "issue": "No panels detected",
                    "description": "Row 1 from column G is empty. Sheet skipped.",
                    "raw_row": None,
                }
            )
            continue

        if not _verify_srno_header(ws):
            actual = ws.cell(row=HEADER_ROW, column=SRNO_COL).value
            logger.warning(
                f"Sheet '{sheet_name}': expected 'SNo' in A4, "
                f"found {repr(actual)!r}. Skipping."
            )
            continue

        header_map = _build_header_map(ws)
        if "DESCRIPTION" not in header_map:
            logger.warning(
                f"Sheet '{sheet_name}': DESCRIPTION header not found "
                "in row 4. Skipping."
            )
            continue

        data_rows: list[dict[str, Any]] = []
        panel_col_indices = {panel_name: col_idx for col_idx, panel_name in panels}

        max_row = ws.max_row
        for row_num in range(HEADER_ROW + 1, max_row + 1):
            raw: dict[str, Any] = {}
            for col_name in STATIC_COLS:
                col_idx = header_map.get(col_name)
                raw[col_name] = (
                    ws.cell(row=row_num, column=col_idx).value if col_idx else None
                )

            raw["panel_quantities"] = dict[str, float]()
            for panel_name, col_idx in panel_col_indices.items():
                cell_val = ws.cell(row=row_num, column=col_idx).value
                raw["panel_quantities"][panel_name] = (
                    float(cell_val) if isinstance(cell_val, (int, float)) else 0.0
                )

            desc_empty = _is_empty(raw.get("DESCRIPTION"))
            cat_empty = _is_empty(raw.get("CAT NO."))
            spec_empty = _is_empty(raw.get("SPEC"))
            make_empty = _is_empty(raw.get("MAKE"))
            unit_empty = _is_empty(raw.get("UNIT"))
            has_qty = any(not _is_empty(v) for v in raw["panel_quantities"].values())

            # Absolute void bypass
            if (
                desc_empty
                and cat_empty
                and spec_empty
                and make_empty
                and unit_empty
                and not has_qty
            ):
                continue

            is_category = (
                not desc_empty
                and cat_empty
                and spec_empty
                and make_empty
                and unit_empty
                and not has_qty
            )

            if is_category:
                current_category = str(raw.get("DESCRIPTION")).strip()
                logger.debug(
                    f"Sheet '{sheet_name}' row {row_num}: "
                    f"category → '{current_category}'"
                )
                continue

            raw["CATEGORY"] = current_category

            ok, cleaned, exc = clean_row(raw, sheet_name, row_num)

            if ok and cleaned is not None:
                data_rows.append(cleaned)
            elif exc is not None:
                exceptions.append(exc)

        logger.info(
            f"Sheet '{sheet_name}': {len(panels)} panel(s), "
            f"{len(data_rows)} valid item(s) extracted."
        )

        sheets_data.append(
            {
                "sheet_name": sheet_name,
                "panels": panels,
                "data_rows": data_rows,
            }
        )

    wb.close()
    return sheets_data, exceptions
